"""
Lobster AI - File Service
Service layer for file operations, metadata analysis, and preview generation.
"""

import pandas as pd
import scanpy as sc
import h5py
import gzip
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
from uuid import UUID
import tempfile
import traceback
from openpyxl import load_workbook

from lobster.utils.file_analyzer import FileAnalyzer, get_workspace_files_metadata
from lobster.utils.logger import get_logger

logger = get_logger(__name__)


class FileService:
    """
    Service class for file operations, metadata analysis, and preview generation.
    Optimized for performance with large bioinformatics datasets.
    """
    
    # Preview limits for performance
    DEFAULT_MAX_ROWS = 10
    DEFAULT_MAX_COLUMNS = 20
    MAX_PREVIEW_ROWS = 100
    MAX_PREVIEW_COLUMNS = 50
    
    # File size limits for preview generation (in bytes)
    MAX_TEXT_FILE_SIZE = 100 * 1024 * 1024  # 100MB for text files
    MAX_BINARY_FILE_SIZE = 1024 * 1024 * 1024  # 1GB for binary files
    
    @classmethod
    def analyze_file_metadata(cls, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Analyze file metadata using the FileAnalyzer utility.
        
        Args:
            file_path: Path to the file to analyze
            
        Returns:
            Dictionary containing comprehensive file metadata
        """
        try:
            return FileAnalyzer.analyze_file_metadata(file_path)
        except Exception as e:
            logger.error(f"Error in file service metadata analysis: {e}")
            raise
    
    @classmethod
    def get_session_files_metadata(cls, session_id: Union[str, UUID], 
                                   workspace_base_path: Union[str, Path] = "workspaces",
                                   directory: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Get metadata for all files in a session workspace.
        
        Args:
            session_id: Session identifier
            workspace_base_path: Base path for workspaces
            directory: Optional subdirectory to search in
            
        Returns:
            List of file metadata dictionaries
        """
        try:
            workspace_path = Path(workspace_base_path) / str(session_id)
            
            if not workspace_path.exists():
                logger.warning(f"Workspace path does not exist: {workspace_path}")
                return []
            
            files_metadata = get_workspace_files_metadata(workspace_path, directory)
            
            # Add session context to each file
            for file_meta in files_metadata:
                file_meta['session_id'] = str(session_id)
            
            logger.info(f"Retrieved metadata for {len(files_metadata)} files in session {session_id}")
            return files_metadata
            
        except Exception as e:
            logger.error(f"Error getting session files metadata for {session_id}: {e}")
            return []
    
    @classmethod
    def generate_file_preview(cls, session_id: Union[str, UUID], 
                             file_path: str,
                             workspace_base_path: Union[str, Path] = "workspaces",
                             max_rows: int = DEFAULT_MAX_ROWS,
                             max_columns: int = DEFAULT_MAX_COLUMNS) -> Dict[str, Any]:
        """
        Generate a preview of file content with specified limits.
        
        Args:
            session_id: Session identifier
            file_path: Relative path to the file within the session workspace
            workspace_base_path: Base path for workspaces
            max_rows: Maximum number of rows to include in preview
            max_columns: Maximum number of columns to include in preview
            
        Returns:
            Dictionary containing preview data and file information
        """
        try:
            # Validate limits
            max_rows = min(max_rows, cls.MAX_PREVIEW_ROWS)
            max_columns = min(max_columns, cls.MAX_PREVIEW_COLUMNS)
            
            # Construct full file path
            workspace_path = Path(workspace_base_path) / str(session_id)
            full_file_path = workspace_path / file_path
            
            if not full_file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Get file metadata first
            file_metadata = cls.analyze_file_metadata(full_file_path)
            
            # Check file size limits
            file_size = file_metadata.get('size_bytes', 0)
            file_format = file_metadata.get('file_format', 'unknown')
            
            if not cls._is_preview_feasible(file_size, file_format):
                return {
                    'success': False,
                    'error': 'File too large for preview generation',
                    'file_info': file_metadata,
                    'preview_data': None
                }
            
            # Generate preview based on file format
            if not file_metadata.get('is_data_file', False):
                return {
                    'success': False,
                    'error': 'File format not supported for preview',
                    'file_info': file_metadata,
                    'preview_data': None
                }
            
            preview_data = cls._generate_preview_by_format(
                full_file_path, file_format, max_rows, max_columns
            )
            
            return {
                'success': True,
                'message': 'Preview generated successfully',
                'file_info': file_metadata,
                'preview_data': preview_data
            }
            
        except FileNotFoundError as e:
            logger.warning(f"File not found for preview: {e}")
            return {
                'success': False,
                'error': str(e),
                'file_info': None,
                'preview_data': None
            }
        except Exception as e:
            logger.error(f"Error generating file preview: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return {
                'success': False,
                'error': f'Preview generation failed: {str(e)}',
                'file_info': None,
                'preview_data': None
            }
    
    @classmethod
    def _is_preview_feasible(cls, file_size: int, file_format: str) -> bool:
        """Check if preview generation is feasible based on file size and format."""
        if file_format in ['csv', 'tsv', 'txt']:
            return file_size <= cls.MAX_TEXT_FILE_SIZE
        elif file_format in ['h5ad', 'h5', 'hdf5', 'xlsx', 'mtx']:
            return file_size <= cls.MAX_BINARY_FILE_SIZE
        elif file_format == 'gz':
            # For compressed files, we're more conservative
            return file_size <= cls.MAX_TEXT_FILE_SIZE // 2
        else:
            return False
    
    @classmethod
    def _generate_preview_by_format(cls, file_path: Path, file_format: str, 
                                   max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview data based on specific file format."""
        try:
            if file_format in ['csv', 'tsv']:
                return cls._preview_text_data(file_path, file_format, max_rows, max_columns)
            elif file_format == 'h5ad':
                return cls._preview_h5ad_data(file_path, max_rows, max_columns)
            elif file_format in ['h5', 'hdf5']:
                return cls._preview_h5_data(file_path, max_rows, max_columns)
            elif file_format == 'xlsx':
                return cls._preview_excel_data(file_path, max_rows, max_columns)
            elif file_format == 'mtx':
                return cls._preview_mtx_data(file_path, max_rows, max_columns)
            elif file_format == 'gz':
                return cls._preview_compressed_data(file_path, max_rows, max_columns)
            else:
                raise ValueError(f"Unsupported format for preview: {file_format}")
                
        except Exception as e:
            logger.error(f"Error generating preview for format {file_format}: {e}")
            raise
    
    @classmethod
    def _preview_text_data(cls, file_path: Path, file_format: str, 
                          max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for CSV/TSV files."""
        delimiter = ',' if file_format == 'csv' else '\t'
        
        try:
            # Use pandas for robust parsing
            df = pd.read_csv(
                file_path, 
                delimiter=delimiter, 
                nrows=max_rows,
                low_memory=False,
                on_bad_lines='skip'
            )
            
            # Limit columns
            if len(df.columns) > max_columns:
                df = df.iloc[:, :max_columns]
                is_truncated_cols = True
            else:
                is_truncated_cols = False
            
            # Convert to preview format
            headers = df.columns.tolist()
            rows = df.values.tolist()
            
            # Convert any non-string values to strings for JSON serialization
            rows = [[str(cell) if pd.notna(cell) else '' for cell in row] for row in rows]
            
            # Get total file dimensions
            total_info = cls._get_text_file_dimensions(file_path, delimiter)
            
            return {
                'headers': headers,
                'rows': rows,
                'total_rows': total_info.get('total_rows', len(rows)),
                'total_columns': total_info.get('total_columns', len(headers)),
                'is_truncated': len(rows) >= max_rows or is_truncated_cols,
                'preview_rows': len(rows),
                'preview_columns': len(headers),
                'delimiter': delimiter
            }
            
        except Exception as e:
            logger.error(f"Error previewing text data {file_path}: {e}")
            raise
    
    @classmethod
    def _preview_h5ad_data(cls, file_path: Path, max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for H5AD files."""
        try:
            # Read with scanpy in backed mode for efficiency
            adata = sc.read_h5ad(file_path, backed='r')
            
            # Get dimensions
            total_rows = adata.n_obs
            total_columns = adata.n_vars
            
            # Limit preview dimensions
            preview_rows = min(max_rows, total_rows)
            preview_cols = min(max_columns, total_columns)
            
            # Extract preview data from the expression matrix
            if hasattr(adata, 'X') and adata.X is not None:
                # Get a subset of the data
                X_subset = adata.X[:preview_rows, :preview_cols]
                
                # Convert to dense if sparse
                if hasattr(X_subset, 'toarray'):
                    X_subset = X_subset.toarray()
                
                # Get observation and variable names
                obs_names = adata.obs_names[:preview_rows].tolist()
                var_names = adata.var_names[:preview_cols].tolist()
                
                # Convert data to list format
                rows = []
                for i, obs_name in enumerate(obs_names):
                    row = [obs_name] + [str(X_subset[i, j]) for j in range(len(var_names))]
                    rows.append(row)
                
                headers = ['obs_names'] + var_names
                
            else:
                # Fallback: use observation metadata if no expression data
                obs_subset = adata.obs.iloc[:preview_rows, :preview_cols]
                headers = obs_subset.columns.tolist()
                rows = obs_subset.values.tolist()
                rows = [[str(cell) if pd.notna(cell) else '' for cell in row] for row in rows]
            
            return {
                'headers': headers,
                'rows': rows,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'is_truncated': preview_rows < total_rows or preview_cols < total_columns,
                'preview_rows': len(rows),
                'preview_columns': len(headers),
                'format_info': {
                    'n_obs': total_rows,
                    'n_vars': total_columns,
                    'obs_keys': list(adata.obs.columns),
                    'var_keys': list(adata.var.columns)
                }
            }
            
        except Exception as e:
            logger.error(f"Error previewing H5AD data {file_path}: {e}")
            raise
    
    @classmethod
    def _preview_h5_data(cls, file_path: Path, max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for HDF5 files."""
        try:
            with h5py.File(file_path, 'r') as f:
                # Find the main dataset
                datasets = []
                
                def collect_datasets(name, obj):
                    if isinstance(obj, h5py.Dataset):
                        datasets.append((name, obj))
                
                f.visititems(collect_datasets)
                
                if not datasets:
                    raise ValueError("No datasets found in HDF5 file")
                
                # Use the largest dataset as main
                main_name, main_dataset = max(datasets, key=lambda x: x[1].size)
                
                # Get dataset shape
                shape = main_dataset.shape
                if len(shape) < 2:
                    # 1D dataset, treat as single column
                    total_rows = shape[0]
                    total_columns = 1
                    
                    preview_rows = min(max_rows, total_rows)
                    data = main_dataset[:preview_rows]
                    
                    headers = [main_name]
                    rows = [[str(val)] for val in data]
                    
                else:
                    # 2D dataset
                    total_rows, total_columns = shape[:2]
                    
                    preview_rows = min(max_rows, total_rows)
                    preview_cols = min(max_columns, total_columns)
                    
                    data = main_dataset[:preview_rows, :preview_cols]
                    
                    headers = [f"col_{i}" for i in range(preview_cols)]
                    rows = [[str(data[i, j]) for j in range(preview_cols)] for i in range(preview_rows)]
                
                return {
                    'headers': headers,
                    'rows': rows,
                    'total_rows': total_rows,
                    'total_columns': total_columns,
                    'is_truncated': len(rows) < total_rows or len(headers) < total_columns,
                    'preview_rows': len(rows),
                    'preview_columns': len(headers),
                    'format_info': {
                        'main_dataset': main_name,
                        'dataset_shape': shape,
                        'available_datasets': [name for name, _ in datasets]
                    }
                }
                
        except Exception as e:
            logger.error(f"Error previewing HDF5 data {file_path}: {e}")
            raise
    
    @classmethod
    def _preview_excel_data(cls, file_path: Path, max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for Excel files."""
        try:
            # Use pandas for robust Excel reading
            df = pd.read_excel(file_path, nrows=max_rows, engine='openpyxl')
            
            # Limit columns
            if len(df.columns) > max_columns:
                df = df.iloc[:, :max_columns]
                is_truncated_cols = True
            else:
                is_truncated_cols = False
            
            headers = df.columns.tolist()
            rows = df.values.tolist()
            
            # Convert to strings for JSON serialization
            rows = [[str(cell) if pd.notna(cell) else '' for cell in row] for row in rows]
            
            # Get total dimensions from workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            total_rows = ws.max_row - 1 if ws.max_row > 1 else 0  # Subtract header
            total_columns = ws.max_column
            
            return {
                'headers': headers,
                'rows': rows,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'is_truncated': len(rows) >= max_rows or is_truncated_cols,
                'preview_rows': len(rows),
                'preview_columns': len(headers),
                'format_info': {
                    'sheet_names': wb.sheetnames,
                    'active_sheet': ws.title
                }
            }
            
        except Exception as e:
            logger.error(f"Error previewing Excel data {file_path}: {e}")
            raise
    
    @classmethod
    def _preview_mtx_data(cls, file_path: Path, max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for Matrix Market files."""
        try:
            from scipy.io import mmread
            
            # Read the sparse matrix
            matrix = mmread(file_path)
            
            # Convert to dense for preview (only small portion)
            if hasattr(matrix, 'toarray'):
                dense_matrix = matrix.toarray()
            else:
                dense_matrix = matrix
            
            total_rows, total_columns = dense_matrix.shape
            
            # Limit preview size
            preview_rows = min(max_rows, total_rows)
            preview_cols = min(max_columns, total_columns)
            
            subset = dense_matrix[:preview_rows, :preview_cols]
            
            headers = [f"col_{i}" for i in range(preview_cols)]
            rows = [[str(subset[i, j]) for j in range(preview_cols)] for i in range(preview_rows)]
            
            return {
                'headers': headers,
                'rows': rows,
                'total_rows': total_rows,
                'total_columns': total_columns,
                'is_truncated': preview_rows < total_rows or preview_cols < total_columns,
                'preview_rows': len(rows),
                'preview_columns': len(headers),
                'format_info': {
                    'matrix_format': 'sparse',
                    'nnz': matrix.nnz if hasattr(matrix, 'nnz') else 'unknown'
                }
            }
            
        except ImportError:
            raise ValueError("scipy required for MTX file preview")
        except Exception as e:
            logger.error(f"Error previewing MTX data {file_path}: {e}")
            raise
    
    @classmethod
    def _preview_compressed_data(cls, file_path: Path, max_rows: int, max_columns: int) -> Dict[str, Any]:
        """Generate preview for compressed files."""
        try:
            # Determine the original format
            stem_path = Path(file_path.stem)
            stem_extension = stem_path.suffix.lower()
            
            if stem_extension == '.csv':
                file_format = 'csv'
            elif stem_extension == '.tsv':
                file_format = 'tsv'
            else:
                file_format = 'txt'
            
            # For text files, we can preview directly from compressed stream
            if file_format in ['csv', 'tsv', 'txt']:
                delimiter = ',' if file_format == 'csv' else '\t'
                
                with gzip.open(file_path, 'rt', encoding='utf-8', errors='ignore') as f:
                    # Read limited rows using pandas
                    df = pd.read_csv(
                        f, 
                        delimiter=delimiter, 
                        nrows=max_rows,
                        low_memory=False,
                        on_bad_lines='skip'
                    )
                    
                    # Limit columns
                    if len(df.columns) > max_columns:
                        df = df.iloc[:, :max_columns]
                        is_truncated_cols = True
                    else:
                        is_truncated_cols = False
                    
                    headers = df.columns.tolist()
                    rows = df.values.tolist()
                    rows = [[str(cell) if pd.notna(cell) else '' for cell in row] for row in rows]
                    
                    # For compressed files, total dimensions are estimates
                    return {
                        'headers': headers,
                        'rows': rows,
                        'total_rows': 'unknown (compressed)',
                        'total_columns': len(headers),
                        'is_truncated': True,  # Always assume truncated for compressed
                        'preview_rows': len(rows),
                        'preview_columns': len(headers),
                        'format_info': {
                            'compressed': True,
                            'original_format': file_format
                        }
                    }
            else:
                raise ValueError(f"Unsupported compressed format: {stem_extension}")
                
        except Exception as e:
            logger.error(f"Error previewing compressed data {file_path}: {e}")
            raise
    
    @classmethod
    def _get_text_file_dimensions(cls, file_path: Path, delimiter: str) -> Dict[str, int]:
        """Get total dimensions of a text file efficiently."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                # Get column count from header
                header_line = f.readline().strip()
                total_columns = len(header_line.split(delimiter)) if header_line else 0
                
                # Count total lines
                f.seek(0)
                total_lines = sum(1 for _ in f)
                
                # Assume first line is header
                total_rows = max(0, total_lines - 1)
                
                return {
                    'total_rows': total_rows,
                    'total_columns': total_columns
                }
        except Exception:
            return {'total_rows': 0, 'total_columns': 0}


# Convenience functions for external usage
def analyze_file_metadata(file_path: Union[str, Path]) -> Dict[str, Any]:
    """Analyze file metadata."""
    return FileService.analyze_file_metadata(file_path)


def get_session_files_metadata(session_id: Union[str, UUID], 
                              workspace_base_path: Union[str, Path] = "workspaces",
                              directory: Optional[str] = None) -> List[Dict[str, Any]]:
    """Get session files metadata."""
    return FileService.get_session_files_metadata(session_id, workspace_base_path, directory)


def generate_file_preview(session_id: Union[str, UUID], 
                         file_path: str,
                         workspace_base_path: Union[str, Path] = "workspaces",
                         max_rows: int = FileService.DEFAULT_MAX_ROWS,
                         max_columns: int = FileService.DEFAULT_MAX_COLUMNS) -> Dict[str, Any]:
    """Generate file preview."""
    return FileService.generate_file_preview(
        session_id, file_path, workspace_base_path, max_rows, max_columns
    )
