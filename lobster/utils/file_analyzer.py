"""
Lobster AI - File Metadata Analyzer
Efficiently extract file metadata without loading full content for performance optimization.
"""

import os
import gzip
import h5py
import pandas as pd
import scanpy as sc
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, Union, Tuple, List
import tempfile
import shutil
from openpyxl import load_workbook

from lobster.utils.logger import get_logger

logger = get_logger(__name__)


class FileAnalyzer:
    """
    Utility class for analyzing file metadata without loading full content.
    Optimized for bioinformatics file formats with performance in mind.
    """
    
    SUPPORTED_FORMATS = {
        '.csv': 'csv',
        '.tsv': 'tsv',
        '.txt': 'txt',
        '.h5ad': 'h5ad',
        '.h5': 'h5',
        '.hdf5': 'hdf5',
        '.xlsx': 'xlsx',
        '.xls': 'xlsx',
        '.gz': 'gz',
        '.mtx': 'mtx'
    }
    
    DATA_FILE_EXTENSIONS = {'.csv', '.tsv', '.txt', '.h5ad', '.h5', '.hdf5', '.xlsx', '.xls', '.mtx'}
    
    @classmethod
    def analyze_file_metadata(cls, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from a file without loading full content.
        
        Args:
            file_path: Path to the file to analyze
            
        Returns:
            Dictionary containing file metadata
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            # Get basic file information
            stat_info = file_path.stat()
            file_extension = file_path.suffix.lower()
            
            # Handle compressed files
            actual_extension = file_extension
            if file_extension == '.gz':
                # Check the extension before .gz
                stem_extension = Path(file_path.stem).suffix.lower()
                if stem_extension in cls.SUPPORTED_FORMATS:
                    actual_extension = stem_extension
            
            # Determine file type and format
            file_format = cls.SUPPORTED_FORMATS.get(actual_extension, 'unknown')
            is_data_file = actual_extension in cls.DATA_FILE_EXTENSIONS or file_extension == '.gz'
            
            metadata = {
                'name': file_path.name,
                'path': str(file_path),
                'size_bytes': stat_info.st_size,
                'created_at': datetime.fromtimestamp(stat_info.st_ctime),
                'modified_at': datetime.fromtimestamp(stat_info.st_mtime),
                'file_type': cls._get_file_type(file_path),
                'file_format': file_format,
                'is_data_file': is_data_file,
                'row_count': None,
                'column_count': None
            }
            
            # Extract data-specific metadata for supported formats
            if is_data_file and stat_info.st_size > 0:
                try:
                    data_info = cls._extract_data_info(file_path, file_format)
                    metadata.update(data_info)
                except Exception as e:
                    logger.warning(f"Could not extract data info from {file_path}: {e}")
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error analyzing file metadata for {file_path}: {e}")
            # Return basic metadata even if data analysis fails
            return {
                'name': file_path.name,
                'path': str(file_path),
                'size_bytes': file_path.stat().st_size if file_path.exists() else 0,
                'created_at': datetime.fromtimestamp(file_path.stat().st_ctime) if file_path.exists() else datetime.now(),
                'modified_at': datetime.fromtimestamp(file_path.stat().st_mtime) if file_path.exists() else datetime.now(),
                'file_type': cls._get_file_type(file_path),
                'file_format': 'unknown',
                'is_data_file': False,
                'row_count': None,
                'column_count': None,
                'error': str(e)
            }
    
    @classmethod
    def _get_file_type(cls, file_path: Path) -> str:
        """Determine the general file type category."""
        extension = file_path.suffix.lower()
        
        if extension == '.gz':
            # Check the extension before .gz
            stem_extension = Path(file_path.stem).suffix.lower()
            extension = stem_extension if stem_extension else extension
        
        if extension in {'.csv', '.tsv', '.txt'}:
            return 'text_data'
        elif extension in {'.h5ad', '.h5', '.hdf5'}:
            return 'hdf5_data'
        elif extension in {'.xlsx', '.xls'}:
            return 'spreadsheet'
        elif extension == '.mtx':
            return 'matrix'
        elif extension in {'.png', '.jpg', '.jpeg', '.svg', '.pdf'}:
            return 'plot'
        else:
            return 'other'
    
    @classmethod
    def _extract_data_info(cls, file_path: Path, file_format: str) -> Dict[str, Any]:
        """
        Extract data-specific information (rows, columns) without loading full content.
        
        Args:
            file_path: Path to the data file
            file_format: Detected file format
            
        Returns:
            Dictionary with row_count, column_count, and additional metadata
        """
        try:
            if file_format in ['csv', 'tsv']:
                return cls._analyze_text_data(file_path, file_format)
            elif file_format == 'h5ad':
                return cls._analyze_h5ad_data(file_path)
            elif file_format in ['h5', 'hdf5']:
                return cls._analyze_h5_data(file_path)
            elif file_format == 'xlsx':
                return cls._analyze_excel_data(file_path)
            elif file_format == 'mtx':
                return cls._analyze_mtx_data(file_path)
            elif file_format == 'gz':
                return cls._analyze_compressed_data(file_path)
            else:
                return {'row_count': None, 'column_count': None}
                
        except Exception as e:
            logger.warning(f"Failed to extract data info from {file_path}: {e}")
            return {'row_count': None, 'column_count': None, 'analysis_error': str(e)}
    
    @classmethod
    def _analyze_text_data(cls, file_path: Path, file_format: str) -> Dict[str, Any]:
        """Analyze CSV/TSV files efficiently by reading only header and sampling."""
        delimiter = ',' if file_format == 'csv' else '\t'
        
        try:
            # Read just the first few lines to get column count and estimate rows
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                # Get header for column count
                header_line = f.readline().strip()
                if header_line:
                    column_count = len(header_line.split(delimiter))
                else:
                    column_count = 0
                
                # Count total lines efficiently
                f.seek(0)
                row_count = sum(1 for _ in f)
                
                # Subtract 1 if there's a header (check if first line looks like header)
                f.seek(0)
                first_line = f.readline().strip()
                second_line = f.readline().strip() if f else ""
                
                has_header = cls._detect_header(first_line, second_line, delimiter)
                if has_header and row_count > 0:
                    row_count -= 1
            
            return {
                'row_count': max(0, row_count),
                'column_count': column_count,
                'has_header': has_header,
                'delimiter': delimiter
            }
            
        except Exception as e:
            logger.warning(f"Error analyzing text data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_h5ad_data(cls, file_path: Path) -> Dict[str, Any]:
        """Analyze H5AD (AnnData) files using scanpy."""
        try:
            # Use scanpy to read just the metadata without loading full data
            adata = sc.read_h5ad(file_path, backed='r')  # Read in backed mode for metadata only
            
            return {
                'row_count': adata.n_obs,
                'column_count': adata.n_vars,
                'format_version': getattr(adata, 'format_version', None),
                'obs_keys': list(adata.obs.columns) if hasattr(adata, 'obs') else [],
                'var_keys': list(adata.var.columns) if hasattr(adata, 'var') else []
            }
            
        except Exception as e:
            logger.warning(f"Error analyzing H5AD data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_h5_data(cls, file_path: Path) -> Dict[str, Any]:
        """Analyze HDF5 files to extract structure information."""
        try:
            with h5py.File(file_path, 'r') as f:
                # Get the main datasets and their shapes
                datasets = []
                total_elements = 0
                
                def collect_datasets(name, obj):
                    if isinstance(obj, h5py.Dataset):
                        datasets.append({
                            'name': name,
                            'shape': obj.shape,
                            'dtype': str(obj.dtype)
                        })
                        nonlocal total_elements
                        total_elements += obj.size
                
                f.visititems(collect_datasets)
                
                # If there's a main dataset, use its dimensions
                main_dataset = None
                if datasets:
                    # Look for common main dataset names or use the largest
                    main_candidates = ['X', 'data', 'matrix']
                    for candidate in main_candidates:
                        for ds in datasets:
                            if candidate in ds['name'].lower():
                                main_dataset = ds
                                break
                        if main_dataset:
                            break
                    
                    if not main_dataset:
                        # Use the largest dataset
                        main_dataset = max(datasets, key=lambda x: x['shape'][0] if x['shape'] else 0)
                
                if main_dataset and main_dataset['shape']:
                    row_count = main_dataset['shape'][0]
                    column_count = main_dataset['shape'][1] if len(main_dataset['shape']) > 1 else 1
                else:
                    row_count = None
                    column_count = None
                
                return {
                    'row_count': row_count,
                    'column_count': column_count,
                    'datasets': [ds['name'] for ds in datasets],
                    'total_datasets': len(datasets),
                    'main_dataset': main_dataset['name'] if main_dataset else None
                }
                
        except Exception as e:
            logger.warning(f"Error analyzing HDF5 data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_excel_data(cls, file_path: Path) -> Dict[str, Any]:
        """Analyze Excel files efficiently."""
        try:
            # Load workbook without data to get structure
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get info from the first sheet
            if wb.worksheets:
                ws = wb.worksheets[0]
                
                # Get max row and column that contain data
                row_count = ws.max_row
                column_count = ws.max_column
                
                # Account for potential empty rows/columns at the end
                if row_count and column_count:
                    # Check if first row looks like header
                    has_header = False
                    if row_count > 1:
                        first_row_values = [cell.value for cell in ws[1]]
                        second_row_values = [cell.value for cell in ws[2]] if row_count > 1 else []
                        has_header = cls._detect_excel_header(first_row_values, second_row_values)
                    
                    if has_header and row_count > 0:
                        row_count -= 1
                else:
                    row_count = 0
                    column_count = 0
                
                return {
                    'row_count': max(0, row_count),
                    'column_count': column_count,
                    'sheet_names': wb.sheetnames,
                    'active_sheet': ws.title,
                    'has_header': has_header
                }
            else:
                return {'row_count': 0, 'column_count': 0}
                
        except Exception as e:
            logger.warning(f"Error analyzing Excel data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_mtx_data(cls, file_path: Path) -> Dict[str, Any]:
        """Analyze Matrix Market (.mtx) files."""
        try:
            with open(file_path, 'r') as f:
                # Skip comments
                line = f.readline()
                while line.startswith('%'):
                    line = f.readline()
                
                # Parse matrix dimensions
                parts = line.strip().split()
                if len(parts) >= 3:
                    rows = int(parts[0])
                    cols = int(parts[1])
                    nnz = int(parts[2])  # number of non-zero entries
                    
                    return {
                        'row_count': rows,
                        'column_count': cols,
                        'non_zero_entries': nnz,
                        'sparsity': (rows * cols - nnz) / (rows * cols) if rows * cols > 0 else 0
                    }
                    
        except Exception as e:
            logger.warning(f"Error analyzing MTX data {file_path}: {e}")
            
        return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_compressed_data(cls, file_path: Path) -> Dict[str, Any]:
        """Analyze compressed (.gz) files by decompressing header."""
        try:
            # Get the original file extension
            stem_path = Path(file_path.stem)
            stem_extension = stem_path.suffix.lower()
            
            if stem_extension in cls.SUPPORTED_FORMATS:
                file_format = cls.SUPPORTED_FORMATS[stem_extension]
                
                # For text files, we can analyze the header without full decompression
                if file_format in ['csv', 'tsv', 'txt']:
                    return cls._analyze_compressed_text_data(file_path, file_format)
                else:
                    # For binary formats, we'd need to decompress temporarily
                    # This is more expensive but still more efficient than loading all data
                    return cls._analyze_compressed_binary_data(file_path, file_format)
            
            return {'row_count': None, 'column_count': None}
            
        except Exception as e:
            logger.warning(f"Error analyzing compressed data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_compressed_text_data(cls, file_path: Path, file_format: str) -> Dict[str, Any]:
        """Analyze compressed text files efficiently."""
        delimiter = ',' if file_format == 'csv' else '\t'
        
        try:
            with gzip.open(file_path, 'rt', encoding='utf-8', errors='ignore') as f:
                # Get header for column count
                header_line = f.readline().strip()
                column_count = len(header_line.split(delimiter)) if header_line else 0
                
                # Count lines efficiently
                f.seek(0)
                row_count = sum(1 for _ in f)
                
                # Check for header
                f.seek(0)
                first_line = f.readline().strip()
                second_line = f.readline().strip() if f else ""
                
                has_header = cls._detect_header(first_line, second_line, delimiter)
                if has_header and row_count > 0:
                    row_count -= 1
                
                return {
                    'row_count': max(0, row_count),
                    'column_count': column_count,
                    'has_header': has_header,
                    'delimiter': delimiter,
                    'compressed': True
                }
                
        except Exception as e:
            logger.warning(f"Error analyzing compressed text data {file_path}: {e}")
            return {'row_count': None, 'column_count': None}
    
    @classmethod
    def _analyze_compressed_binary_data(cls, file_path: Path, file_format: str) -> Dict[str, Any]:
        """Analyze compressed binary files by temporary decompression of header."""
        try:
            # Create a temporary file for partial decompression
            with tempfile.NamedTemporaryFile(suffix=f'.{file_format}', delete=False) as temp_file:
                temp_path = Path(temp_file.name)
                
                # Decompress just enough to analyze structure
                with gzip.open(file_path, 'rb') as gz_file:
                    # Copy a reasonable amount for analysis (e.g., first 1MB for structure)
                    chunk_size = 1024 * 1024  # 1MB
                    data = gz_file.read(chunk_size)
                    temp_file.write(data)
                
                # Analyze the temporary file
                try:
                    if file_format == 'h5ad':
                        result = cls._analyze_h5ad_data(temp_path)
                    elif file_format in ['h5', 'hdf5']:
                        result = cls._analyze_h5_data(temp_path)
                    else:
                        result = {'row_count': None, 'column_count': None}
                    
                    result['compressed'] = True
                    return result
                    
                finally:
                    # Clean up temporary file
                    if temp_path.exists():
                        temp_path.unlink()
                        
        except Exception as e:
            logger.warning(f"Error analyzing compressed binary data {file_path}: {e}")
            return {'row_count': None, 'column_count': None, 'compressed': True}
    
    @classmethod
    def _detect_header(cls, first_line: str, second_line: str, delimiter: str) -> bool:
        """Detect if the first line is a header based on content analysis."""
        if not first_line or not second_line:
            return True  # Assume header if we can't compare
        
        try:
            first_values = first_line.split(delimiter)
            second_values = second_line.split(delimiter)
            
            if len(first_values) != len(second_values):
                return True  # Different column counts suggest header
            
            # Check if first row has more text and second row has more numbers
            first_numeric = sum(1 for val in first_values if cls._is_numeric(val.strip()))
            second_numeric = sum(1 for val in second_values if cls._is_numeric(val.strip()))
            
            # If second row has significantly more numeric values, first is likely header
            if len(first_values) > 0:
                first_numeric_ratio = first_numeric / len(first_values)
                second_numeric_ratio = second_numeric / len(second_values)
                
                return second_numeric_ratio > first_numeric_ratio + 0.3
            
            return True
            
        except Exception:
            return True  # Default to assuming header
    
    @classmethod
    def _detect_excel_header(cls, first_row_values: list, second_row_values: list) -> bool:
        """Detect if the first row in Excel is a header."""
        if not first_row_values or not second_row_values:
            return True
        
        try:
            first_numeric = sum(1 for val in first_row_values if val is not None and cls._is_numeric(str(val)))
            second_numeric = sum(1 for val in second_row_values if val is not None and cls._is_numeric(str(val)))
            
            if len(first_row_values) > 0:
                first_numeric_ratio = first_numeric / len(first_row_values)
                second_numeric_ratio = second_numeric / len(second_row_values) if second_row_values else 0
                
                return second_numeric_ratio > first_numeric_ratio + 0.3
            
            return True
            
        except Exception:
            return True
    
    @classmethod
    def _is_numeric(cls, value: str) -> bool:
        """Check if a string value represents a number."""
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False


def analyze_file_metadata(file_path: Union[str, Path]) -> Dict[str, Any]:
    """
    Convenience function to analyze file metadata.
    
    Args:
        file_path: Path to the file to analyze
        
    Returns:
        Dictionary containing file metadata
    """
    return FileAnalyzer.analyze_file_metadata(file_path)


def get_workspace_files_metadata(workspace_path: Union[str, Path], 
                                 directory: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Get metadata for all files in a workspace directory.
    
    Args:
        workspace_path: Path to the workspace root
        directory: Optional subdirectory to search in
        
    Returns:
        List of file metadata dictionaries
    """
    workspace_path = Path(workspace_path)
    
    if directory:
        search_path = workspace_path / directory
    else:
        search_path = workspace_path
    
    files_metadata = []
    
    try:
        if search_path.exists():
            for file_path in search_path.rglob("*"):
                if file_path.is_file():
                    try:
                        metadata = FileAnalyzer.analyze_file_metadata(file_path)
                        # Add relative path information
                        metadata['relative_path'] = str(file_path.relative_to(workspace_path))
                        metadata['directory'] = str(file_path.parent.relative_to(workspace_path))
                        files_metadata.append(metadata)
                    except Exception as e:
                        logger.warning(f"Could not analyze file {file_path}: {e}")
        
    except Exception as e:
        logger.error(f"Error listing workspace files: {e}")
    
    return files_metadata
